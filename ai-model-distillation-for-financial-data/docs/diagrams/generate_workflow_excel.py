"""Generate workflow sequence diagram Excel export."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Style definitions ──
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

WORKFLOW_FILLS = {
    1: PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid"),  # blue
    2: PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid"),  # green
    3: PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid"),  # red
    4: PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid"),  # orange
    5: PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid"),  # purple
    6: PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),  # teal
}


def style_header(ws, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CENTER
        cell.border = THIN_BORDER


def style_row(ws, row, cols, fill=None, font=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font or BODY_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def auto_width(ws, cols, min_w=12, max_w=55):
    for c in range(1, cols + 1):
        best = min_w
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            val_str = str(row[0].value or "")
            best = max(best, min(max_w, len(val_str) + 4))
        ws.column_dimensions[get_column_letter(c)].width = best


# ════════════════════════════════════════════════════
# Sheet 1: Sequence Steps
# ════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Sequence Steps"
ws1.sheet_properties.tabColor = "2F5496"

headers = ["Step", "Workflow", "Stage", "From", "To", "Action / Message", "Data Exchanged", "Type"]
for c, h in enumerate(headers, 1):
    ws1.cell(row=1, column=c, value=h)
style_header(ws1, len(headers))

rows = [
    # ── Workflow 1: Job Creation ──
    (1,  "1 - Job Creation", "",  "User",         "FastAPI",        "POST /api/jobs",                              "workload_id, client_id",                "HTTP Request"),
    (2,  "1 - Job Creation", "",  "FastAPI",       "FastAPI",        "Build FlywheelRun(status=PENDING)",           "",                                      "Internal"),
    (3,  "1 - Job Creation", "",  "FastAPI",       "KDBXCollection", "db.flywheel_runs.insert_one(run)",            "FlywheelRun document",                  "DB Write"),
    (4,  "1 - Job Creation", "",  "KDBXCollection","KDB-X",          "flywheel_runs insert col1 col2!(v0;v1)",      "Row data via PyKX IPC",                 "q IPC"),
    (5,  "1 - Job Creation", "",  "KDB-X",         "KDBXCollection", "Return inserted_id",                          "ObjectId hex string",                   "q IPC"),
    (6,  "1 - Job Creation", "",  "FastAPI",       "Redis",          "run_nim_workflow_dag.delay(...)",              "workload_id, run_id, client_id, split_config", "Celery Dispatch"),
    (7,  "1 - Job Creation", "",  "FastAPI",       "User",           "Return JobResponse",                          "id, status: queued",                    "HTTP Response"),

    # ── Workflow 2: Flywheel Pipeline ──
    # Stage 1
    (8,  "2 - Flywheel Pipeline", "1 - Initialize",  "Redis",         "Parent Worker", "Dequeue run_nim_workflow_dag",        "",                          "Celery"),
    (9,  "2 - Flywheel Pipeline", "1 - Initialize",  "Parent Worker", "Celery Worker", "initialize_workflow.delay()",          "TaskResult",                "Celery"),
    (10, "2 - Flywheel Pipeline", "1 - Initialize",  "Celery Worker", "TaskDBManager", "check_cancellation(run_id)",           "flywheel_run_id",           "Internal"),
    (11, "2 - Flywheel Pipeline", "1 - Initialize",  "TaskDBManager", "KDB-X",         "select from flywheel_runs where ...",  "run_id filter",             "q IPC"),
    (12, "2 - Flywheel Pipeline", "1 - Initialize",  "Celery Worker", "TaskDBManager", "Update status to RUNNING",            "status=RUNNING",            "DB Write"),
    (13, "2 - Flywheel Pipeline", "1 - Initialize",  "Celery Worker", "TaskDBManager", "Insert NIMRun per NIM in config",     "NIMRun(status=PENDING) x N","DB Write"),
    (14, "2 - Flywheel Pipeline", "1 - Initialize",  "Celery Worker", "Parent Worker", "Return TaskResult",                   "workload_id, run_id, nims", "Celery"),

    # Stage 2
    (15, "2 - Flywheel Pipeline", "2 - Create Datasets", "Parent Worker", "Celery Worker", "create_datasets.delay()",           "TaskResult",                         "Celery"),
    (16, "2 - Flywheel Pipeline", "2 - Create Datasets", "Celery Worker", "KDB-X",         "Read training records from flywheel_logs", "client_id, workload_id filter", "q IPC"),
    (17, "2 - Flywheel Pipeline", "2 - Create Datasets", "Celery Worker", "KDB-X",         "[Optional] enrich_training_pairs_batch()", "Join with market_ticks table",  "q IPC"),
    (18, "2 - Flywheel Pipeline", "2 - Create Datasets", "Celery Worker", "Celery Worker", "identify_workload_type(), split train/eval sets", "",                      "Internal"),
    (19, "2 - Flywheel Pipeline", "2 - Create Datasets", "Celery Worker", "NeMo Datastore","DataUploader.upload(train_dataset)",  "JSONL training data",               "HTTP POST"),
    (20, "2 - Flywheel Pipeline", "2 - Create Datasets", "Celery Worker", "NeMo Datastore","DataUploader.upload(eval_dataset)",   "JSONL data",                        "HTTP POST"),
    (21, "2 - Flywheel Pipeline", "2 - Create Datasets", "Celery Worker", "TaskDBManager", "Update flywheel_runs.datasets",       "dataset names + nmp_uris",          "DB Write"),
    (22, "2 - Flywheel Pipeline", "2 - Create Datasets", "Celery Worker", "Parent Worker", "Return TaskResult",                   "datasets, workload_type",           "Celery"),

    # Stage 3
    (23, "2 - Flywheel Pipeline", "3 - LLM Judge",    "Parent Worker", "Celery Worker", "wait_for_llm_as_judge.delay()",     "TaskResult",                      "Celery"),
    (24, "2 - Flywheel Pipeline", "3 - LLM Judge",    "Celery Worker", "Celery Worker", "[Remote] Validate endpoint reachable", "",                             "Internal"),
    (25, "2 - Flywheel Pipeline", "3 - LLM Judge",    "Celery Worker", "NeMo DMS",      "[Local] deploy_model(judge_nim)",   "NIM config for judge model",      "HTTP POST"),
    (26, "2 - Flywheel Pipeline", "3 - LLM Judge",    "Celery Worker", "NeMo DMS",      "Poll get_deployment_status() until READY", "",                         "HTTP GET (loop)"),
    (27, "2 - Flywheel Pipeline", "3 - LLM Judge",    "Celery Worker", "NIM",           "Poll GET /v1/models until synced",  "",                                "HTTP GET (loop)"),
    (28, "2 - Flywheel Pipeline", "3 - LLM Judge",    "Celery Worker", "TaskDBManager", "Update llm_judge_runs.deployment_status", "deployment_status",          "DB Write"),

    # Stage 4a
    (29, "2 - Flywheel Pipeline", "4a - Spin Up NIM",  "Parent Worker", "Celery Worker", "spin_up_nim.delay(nim_config)",     "NIMConfig",                       "Celery"),
    (30, "2 - Flywheel Pipeline", "4a - Spin Up NIM",  "Celery Worker", "NeMo DMS",      "POST /v1/deployment/model-deployments", "Model image, GPU count, PVC size", "HTTP POST"),
    (31, "2 - Flywheel Pipeline", "4a - Spin Up NIM",  "Celery Worker", "NeMo DMS",      "Poll deployment status every 5s",   "Checks cancellation each iteration","HTTP GET (loop)"),
    (32, "2 - Flywheel Pipeline", "4a - Spin Up NIM",  "Celery Worker", "NIM",           "Poll GET /v1/models every 30s",     "Wait for model to appear",        "HTTP GET (loop)"),
    (33, "2 - Flywheel Pipeline", "4a - Spin Up NIM",  "Celery Worker", "TaskDBManager", "Update nims.status = RUNNING",      "deployment_status=READY",         "DB Write"),

    # Stage 4b - Base Eval (sequential)
    (34, "2 - Flywheel Pipeline", "4b - Base Eval",              "Parent Worker", "Celery Worker", "run_base_eval.delay()",                 "TaskResult",                    "Celery"),
    (35, "2 - Flywheel Pipeline", "4b - Base Eval",              "Celery Worker", "TaskDBManager", "Insert NIMEvaluation(type=BASE)",        "NIMEvaluation record",          "DB Write"),
    (36, "2 - Flywheel Pipeline", "4b - Base Eval",              "Celery Worker", "NeMo Evaluator","POST /v1/evaluation/jobs",               "dataset, model, metrics config", "HTTP POST"),
    (37, "2 - Flywheel Pipeline", "4b - Base Eval",              "Celery Worker", "NeMo Evaluator","Poll status every 5s until COMPLETED",   "Updates progress in KDB-X",     "HTTP GET (loop)"),
    (38, "2 - Flywheel Pipeline", "4b - Base Eval",              "Celery Worker", "NeMo Evaluator","GET /v1/evaluation/jobs/id/results",     "f1_score or function_accuracy", "HTTP GET"),
    (39, "2 - Flywheel Pipeline", "4b - Base Eval",              "Celery Worker", "MLflow",        "[Optional] upload_evaluation_results()", "Scores + artifacts",            "HTTP POST"),
    (40, "2 - Flywheel Pipeline", "4b - Base Eval",              "Celery Worker", "TaskDBManager", "Update evaluations.scores",              "Final scores",                  "DB Write"),

    # Stage 4c - Generate Signals (base)
    (41, "2 - Flywheel Pipeline", "4c - Generate Signals (base)","Parent Worker", "Celery Worker", "generate_signals.delay(model_type='base')", "TaskResult",                "Celery"),
    (42, "2 - Flywheel Pipeline", "4c - Generate Signals (base)","Celery Worker", "KDB-X",         "Get eval records via RecordExporter",     "workload_id, client_id",        "q IPC"),
    (43, "2 - Flywheel Pipeline", "4c - Generate Signals (base)","Celery Worker", "NIM",           "POST /v1/chat/completions per record",   "request.messages",              "HTTP POST (loop)"),
    (44, "2 - Flywheel Pipeline", "4c - Generate Signals (base)","Celery Worker", "KDB-X",         "write_signals_batch()",                  "BUY/SELL/HOLD signals",         "q IPC"),

    # Stage 4d - Backtest (base)
    (45, "2 - Flywheel Pipeline", "4d - Backtest (base)",        "Parent Worker", "Celery Worker", "run_backtest_assessment.delay(model_type='base')", "TaskResult",          "Celery"),
    (46, "2 - Flywheel Pipeline", "4d - Backtest (base)",        "Celery Worker", "KDB-X",         "Count signals for model_id",             "model_id filter",               "q IPC"),
    (47, "2 - Flywheel Pipeline", "4d - Backtest (base)",        "Celery Worker", "KDB-X",         "run_backtest(model_id, cost_bps)",       "Sharpe, drawdown, return, win_rate", "q IPC"),
    (48, "2 - Flywheel Pipeline", "4d - Backtest (base)",        "Celery Worker", "TaskDBManager", "Update evaluations.scores (backtest)",   "Backtest result scores",        "DB Write"),

    # Stage 4e - Customization
    (49, "2 - Flywheel Pipeline", "4e - Customization",          "Parent Worker", "Celery Worker", "start_customization.delay()",            "TaskResult",                    "Celery"),
    (50, "2 - Flywheel Pipeline", "4e - Customization",          "Celery Worker", "TaskDBManager", "Insert NIMCustomization",                "NIMCustomization record",       "DB Write"),
    (51, "2 - Flywheel Pipeline", "4e - Customization",          "Celery Worker", "NeMo Customizer","start_training_job()",                  "train_dataset, LoRA config, epochs", "HTTP POST"),
    (52, "2 - Flywheel Pipeline", "4e - Customization",          "Celery Worker", "NeMo Customizer","Poll job status until COMPLETED",       "epochs_completed, steps_completed",  "HTTP GET (loop)"),
    (53, "2 - Flywheel Pipeline", "4e - Customization",          "Celery Worker", "NIM",           "Poll GET /v1/models until fine-tuned model synced", "",               "HTTP GET (loop)"),
    (54, "2 - Flywheel Pipeline", "4e - Customization",          "Celery Worker", "TaskDBManager", "Update customizations.customized_model", "Fine-tuned model name",         "DB Write"),

    # Stage 4f - Post-customization eval
    (55, "2 - Flywheel Pipeline", "4f - Customization Eval",     "Parent Worker", "Celery Worker", "run_customization_eval.delay()",         "TaskResult",                    "Celery"),
    (56, "2 - Flywheel Pipeline", "4f - Customization Eval",     "Celery Worker", "TaskDBManager", "Insert NIMEvaluation(type=CUSTOMIZED)",  "NIMEvaluation record",          "DB Write"),
    (57, "2 - Flywheel Pipeline", "4f - Customization Eval",     "Celery Worker", "NeMo Evaluator","POST /v1/evaluation/jobs (fine-tuned)",  "dataset, fine-tuned model",     "HTTP POST"),
    (58, "2 - Flywheel Pipeline", "4f - Customization Eval",     "Celery Worker", "NeMo Evaluator","Poll and get results",                   "Fine-tuned model scores",       "HTTP GET (loop)"),
    (59, "2 - Flywheel Pipeline", "4f - Customization Eval",     "Celery Worker", "MLflow",        "[Optional] upload_evaluation_results()",  "Scores + artifacts",           "HTTP POST"),

    # Stage 4g - Generate Signals (customized)
    (60, "2 - Flywheel Pipeline", "4g - Generate Signals (cust)","Parent Worker", "Celery Worker", "generate_signals.delay(model_type='customized')", "TaskResult",          "Celery"),
    (61, "2 - Flywheel Pipeline", "4g - Generate Signals (cust)","Celery Worker", "KDB-X",         "Get eval records via RecordExporter",     "workload_id, client_id",       "q IPC"),
    (62, "2 - Flywheel Pipeline", "4g - Generate Signals (cust)","Celery Worker", "NIM",           "POST /v1/chat/completions per record",   "request.messages",              "HTTP POST (loop)"),
    (63, "2 - Flywheel Pipeline", "4g - Generate Signals (cust)","Celery Worker", "KDB-X",         "write_signals_batch()",                  "BUY/SELL/HOLD signals",         "q IPC"),

    # Stage 4h - Backtest (customized)
    (64, "2 - Flywheel Pipeline", "4h - Backtest (customized)",  "Parent Worker", "Celery Worker", "run_backtest_assessment.delay(model_type='customized')", "TaskResult",  "Celery"),
    (65, "2 - Flywheel Pipeline", "4h - Backtest (customized)",  "Celery Worker", "KDB-X",         "Count signals for model_id",             "model_id filter",               "q IPC"),
    (66, "2 - Flywheel Pipeline", "4h - Backtest (customized)",  "Celery Worker", "KDB-X",         "run_backtest(model_id, cost_bps)",       "Sharpe, drawdown, return, win_rate", "q IPC"),
    (67, "2 - Flywheel Pipeline", "4h - Backtest (customized)",  "Celery Worker", "TaskDBManager", "Update evaluations.scores (backtest)",   "Backtest result scores",        "DB Write"),

    # Stage 4i - shutdown
    (68, "2 - Flywheel Pipeline", "4i - Shutdown NIM", "Parent Worker", "Celery Worker", "shutdown_deployment.delay()",           "TaskResult",                    "Celery"),
    (69, "2 - Flywheel Pipeline", "4i - Shutdown NIM", "Celery Worker", "TaskDBManager", "Update nims.status = COMPLETED",       "status=COMPLETED",              "DB Write"),
    (70, "2 - Flywheel Pipeline", "4i - Shutdown NIM", "Celery Worker", "NeMo DMS",      "DELETE /v1/deployment/model-deployments","namespace/model",              "HTTP DELETE"),

    # Stage 5 - finalize
    (71, "2 - Flywheel Pipeline", "5 - Finalize",     "Parent Worker", "Celery Worker", "finalize_flywheel_run.delay()",         "TaskResult",                    "Celery"),
    (72, "2 - Flywheel Pipeline", "5 - Finalize",     "Celery Worker", "Celery Worker", "sleep(60s) - allow k8s pod cleanup",    "",                              "Internal"),
    (73, "2 - Flywheel Pipeline", "5 - Finalize",     "Celery Worker", "TaskDBManager", "mark_flywheel_run_completed()",         "status=COMPLETED, finished_at", "DB Write"),
    (74, "2 - Flywheel Pipeline", "5 - Finalize",     "TaskDBManager", "KDB-X",         "update from flywheel_runs where ...",   "Set status + timestamp",        "q IPC"),

    # ── Workflow 3: Cancellation ──
    (75, "3 - Job Cancellation", "", "User",         "FastAPI",        "POST /api/jobs/id/cancel",                  "job_id",                                "HTTP Request"),
    (76, "3 - Job Cancellation", "", "FastAPI",       "JobService",     "cancel_job(id)",                            "job_id",                                "Internal"),
    (77, "3 - Job Cancellation", "", "JobService",    "TaskDBManager",  "get_flywheel_run(id) - validate cancelable","Checks finished_at=None",               "DB Read"),
    (78, "3 - Job Cancellation", "", "JobService",    "Redis",          "cancel_job_resources.delay(id)",            "job_id",                                "Celery Dispatch"),
    (79, "3 - Job Cancellation", "", "FastAPI",       "User",           "Return JobCancelResponse",                  "status: cancelling",                    "HTTP Response"),
    (80, "3 - Job Cancellation", "", "Redis",         "Celery Worker",  "Dequeue cancel_job_resources",              "job_id",                                "Celery"),
    (81, "3 - Job Cancellation", "", "Celery Worker", "TaskDBManager",  "mark_flywheel_run_cancelled()",             "status=CANCELLED, finished_at=now()",   "DB Write"),
    (82, "3 - Job Cancellation", "", "TaskDBManager", "KDB-X",          "update from flywheel_runs where ...",       "Set cancelled flag",                    "q IPC"),
    (83, "3 - Job Cancellation", "", "",              "",               "Running tasks detect via check_cancellation() on next poll", "Database flag pattern", "Distributed"),

    # ── Workflow 4: Deletion ──
    (84, "4 - Job Deletion", "", "User",         "FastAPI",        "DELETE /api/jobs/id",                   "job_id",                                "HTTP Request"),
    (85, "4 - Job Deletion", "", "FastAPI",       "JobService",     "delete_job(id) - validate finished",   "Checks finished_at is not None",        "Internal"),
    (86, "4 - Job Deletion", "", "JobService",    "Redis",          "delete_job_resources.delay(id)",        "job_id",                                "Celery Dispatch"),
    (87, "4 - Job Deletion", "", "FastAPI",       "User",           "Return JobDeleteResponse",              "status: deleting",                      "HTTP Response"),
    (88, "4 - Job Deletion", "", "Celery Worker", "NeMo Customizer","delete_customized_model() per NIM",    "model_name",                            "HTTP DELETE"),
    (89, "4 - Job Deletion", "", "Celery Worker", "NeMo Evaluator", "delete_evaluation_job() per record",   "job_id",                                "HTTP DELETE"),
    (90, "4 - Job Deletion", "", "Celery Worker", "MLflow",         "[Optional] delete_experiment()",        "experiment_name",                       "HTTP DELETE"),
    (91, "4 - Job Deletion", "", "Celery Worker", "NeMo Datastore", "delete + unregister per dataset",      "dataset_name",                          "HTTP DELETE"),
    (92, "4 - Job Deletion", "", "Celery Worker", "KDB-X",          "Cascade delete: evaluations, customizations, nims, llm_judge_runs, flywheel_runs", "All records by job_id", "q IPC"),

    # ── Workflow 5: Vector Search ──
    (94,  "5 - Vector Search", "", "Celery Worker", "HNSW Adapter",  "index_embeddings_to_es(index, vecs, records)", "Embedding vectors + metadata",   "Internal"),
    (95,  "5 - Vector Search", "", "HNSW Adapter",  "KDB-X",         "Insert into flywheel_embeddings table",  "doc_id, embedding, tool_name, record",  "q IPC"),
    (96,  "5 - Vector Search", "", "HNSW Adapter",  "KDB-X",         ".ai.hnsw.put - build native HNSW index", "Vectors, cosine similarity, M=32",     "q IPC"),
    (97,  "5 - Vector Search", "", "Celery Worker", "HNSW Adapter",  "search_similar_embeddings(index, query, k)", "Query vector, top-k",              "Internal"),
    (98,  "5 - Vector Search", "", "HNSW Adapter",  "KDB-X",         ".ai.hnsw.search cosine similarity top-k", "Query vector, k, index ref",         "q IPC"),
    (99,  "5 - Vector Search", "", "KDB-X",         "HNSW Adapter",  "Return (scores, indices)",               "Float scores + row indices",            "q IPC"),
    (100, "5 - Vector Search", "", "HNSW Adapter",  "KDB-X",         "Select records by indices",              "Row indices",                           "q IPC"),
    (101, "5 - Vector Search", "", "HNSW Adapter",  "Celery Worker", "Return (score, tool_name, record) tuples","Ranked results",                      "Internal"),

    # ── Workflow 6: Data Explorer ──
    (102, "6 - Data Explorer", "", "User",    "FastAPI", "GET /api/data/{table}?limit=50",       "table name, limit",               "HTTP Request"),
    (103, "6 - Data Explorer", "", "FastAPI", "KDB-X",   "select[50] from table",                "Parameterized q query",           "q IPC"),
    (104, "6 - Data Explorer", "", "KDB-X",  "FastAPI",  "Return rows",                          "PyKX table -> dicts",             "q IPC"),
    (105, "6 - Data Explorer", "", "FastAPI", "User",     "Return data + total",                  "JSON records + count",            "HTTP Response"),
    (106, "6 - Data Explorer", "", "User",    "FastAPI",  "GET /api/data/schema",                 "",                                "HTTP Request"),
    (107, "6 - Data Explorer", "", "FastAPI", "KDB-X",    "meta each table",                      "Table introspection",             "q IPC"),
    (108, "6 - Data Explorer", "", "FastAPI", "User",     "Return table schemas",                 "Column names + types per table",  "HTTP Response"),
    (109, "6 - Data Explorer", "", "User",    "FastAPI",  "POST /api/backtest",                   "model_id, cost_bps",              "HTTP Request"),
    (110, "6 - Data Explorer", "", "FastAPI", "KDB-X",    "run_backtest(model_id, cost_bps)",     "q analytics on signals table",    "q IPC"),
    (111, "6 - Data Explorer", "", "FastAPI", "User",     "Return BacktestResponse",              "sharpe, max_drawdown, total_return, win_rate", "HTTP Response"),
]

for i, row in enumerate(rows, 2):
    for c, val in enumerate(row, 1):
        ws1.cell(row=i, column=c, value=val)
    wf_num = int(row[1][0])
    fill = WORKFLOW_FILLS.get(wf_num)
    style_row(ws1, i, len(headers), fill=fill)

auto_width(ws1, len(headers))
ws1.auto_filter.ref = "A1:H{}".format(len(rows) + 1)
ws1.freeze_panes = "A2"


# ════════════════════════════════════════════════════
# Sheet 2: Component Reference
# ════════════════════════════════════════════════════
ws2 = wb.create_sheet("Components")
ws2.sheet_properties.tabColor = "548235"

headers2 = ["Component", "Type", "File", "Description", "External Service"]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=c, value=h)
style_header(ws2, len(headers2))

components = [
    ("FastAPI (endpoints.py)", "HTTP Layer",        "src/api/endpoints.py",              "REST API: jobs CRUD, data explorer, backtest, health check", ""),
    ("JobService",            "Service Layer",      "src/api/job_service.py",            "Job lifecycle: create, cancel, delete, detail views",         ""),
    ("Parent Worker",         "Celery (parent_queue)","src/tasks/tasks.py",              "DAG orchestrator, serializes flywheel runs (concurrency=1)",  "Redis"),
    ("Celery Worker",         "Celery (default)",   "src/tasks/tasks.py",                "Executes pipeline stages: init, datasets, eval, customize",   "Redis"),
    ("TaskDBManager",         "DB Facade",          "src/api/db_manager.py",             "Single DB interface for all tasks - wraps KDBXDatabase",      ""),
    ("KDBXDatabase",          "Compat Shim",        "kdbx/compat.py",                    "pymongo-compatible API translating to parameterized q",        ""),
    ("KDBXCollection",        "Compat Shim",        "kdbx/compat.py",                    "find, find_one, insert_one, update_one, delete_many -> q",    ""),
    ("pykx_connection()",     "Connection",         "kdbx/connection.py",                "Context manager for SyncQConnection IPC to KDB-X",            "KDB-X"),
    ("KDB-X",                 "Database",           "(external)",                        "Unified store: 7 tables, HNSW vector index, q analytics",     "KDB-X (port 8082)"),
    ("HNSW Adapter",          "Vector Search",      "kdbx/es_adapter.py",                "Elasticsearch replacement: native HNSW via .ai module",       "KDB-X"),
    ("Schema",                "DDL",                "kdbx/schema.py",                    "7 table definitions using flip syntax",                        "KDB-X"),
    ("DMSClient",             "NeMo Integration",   "src/lib/nemo/dms_client.py",        "Deploy/shutdown NIM containers via NeMo DMS REST API",        "NeMo DMS"),
    ("Evaluator",             "NeMo Integration",   "src/lib/nemo/evaluator.py",         "Submit, poll, and fetch evaluation job results",               "NeMo Evaluator"),
    ("Customizer",            "NeMo Integration",   "src/lib/nemo/customizer.py",        "Submit LoRA SFT training, poll status, wait for model sync",  "NeMo Customizer"),
    ("DataUploader",          "NeMo Integration",   "src/lib/nemo/data_uploader.py",     "Upload/delete JSONL datasets to NeMo Datastore (HF API)",    "NeMo Datastore"),
    ("LLMAsJudge",            "NeMo Integration",   "src/lib/nemo/llm_as_judge.py",      "Validate and manage LLM judge (local or remote)",             "NIM / NVIDIA API"),
    ("FlywheelJobManager",    "Business Logic",     "src/lib/flywheel/job_manager.py",   "Soft cancel (DB flag) and hard delete (cascade cleanup)",     ""),
    ("CleanupManager",        "Business Logic",     "src/lib/flywheel/cleanup_manager.py","Worker shutdown: cancel all in-flight runs + NeMo resources",""),
    ("RecordExporter",        "Business Logic",     "src/lib/flywheel/job_manager.py",   "Read training records from flywheel_logs, apply split config",""),
    ("MLflow",                "Tracking",           "(external)",                        "Experiment tracking, evaluation artifact storage",             "MLflow (port 5000)"),
    ("Redis",                 "Broker",             "(external)",                        "Celery task broker and result backend",                        "Redis (port 6379)"),
    ("NIM",                   "Inference",          "(external)",                        "Model inference endpoint, model listing for sync checks",      "NIM (NeMo proxy)"),
]

for i, row in enumerate(components, 2):
    for c, val in enumerate(row, 1):
        ws2.cell(row=i, column=c, value=val)
    style_row(ws2, i, len(headers2))

auto_width(ws2, len(headers2))
ws2.auto_filter.ref = "A1:E{}".format(len(components) + 1)
ws2.freeze_panes = "A2"


# ════════════════════════════════════════════════════
# Sheet 3: KDB-X Tables
# ════════════════════════════════════════════════════
ws3 = wb.create_sheet("KDB-X Tables")
ws3.sheet_properties.tabColor = "BF8F00"

headers3 = ["Table", "Purpose", "Key Columns", "Written By", "Read By"]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=c, value=h)
style_header(ws3, len(headers3))

tables = [
    ("flywheel_runs",       "Job lifecycle tracking",          "_id, workload_id, client_id, status, started_at, finished_at, datasets (JSON), enrichment_stats (JSON), error",
     "initialize_workflow, create_datasets, finalize, cancel",  "All tasks (cancellation check), API endpoints"),
    ("nims",                "NIM deployment tracking per job",  "_id, flywheel_run_id, model_name, status, deployment_status, runtime_seconds",
     "initialize_workflow, spin_up_nim, shutdown_deployment",    "spin_up_nim, eval tasks, shutdown_deployment"),
    ("evaluations",         "Evaluation job results",           "_id, nim_id, flywheel_run_id, eval_type (BASE/CUSTOMIZED/BACKTEST), scores (JSON), progress, nmp_uri, mlflow_uri",
     "run_base_eval, run_customization_eval, run_backtest",     "API job detail endpoint"),
    ("customizations",      "Fine-tune job tracking",           "_id, nim_id, workload_id, base_model, customized_model, epochs_completed, steps_completed, progress",
     "start_customization",                                     "run_customization_eval, shutdown_deployment, delete_job"),
    ("llm_judge_runs",      "LLM judge deployment status",      "_id, flywheel_run_id, model_name, deployment_status",
     "initialize_workflow, wait_for_llm_as_judge",              "shutdown_deployment (avoid duplicate shutdown)"),
    ("flywheel_logs",       "Request/response audit log",       "doc_id, workload_id, client_id, timestamp, request (JSON), response (JSON)",
     "External ingest (API logging)",                           "create_datasets (training data source)"),
    ("flywheel_embeddings", "Embedding vectors + HNSW index",   "doc_id, index_name, embedding (float[]), tool_name, query_text, record (JSON)",
     "index_embeddings_to_es()",                                "search_similar_embeddings(), _rebuild_hnsw_from_table()"),
]

for i, row in enumerate(tables, 2):
    for c, val in enumerate(row, 1):
        ws3.cell(row=i, column=c, value=val)
    style_row(ws3, i, len(headers3))

auto_width(ws3, len(headers3))
ws3.freeze_panes = "A2"


# ════════════════════════════════════════════════════
# Sheet 4: Status Transitions
# ════════════════════════════════════════════════════
ws4 = wb.create_sheet("Status Transitions")
ws4.sheet_properties.tabColor = "C00000"

headers4 = ["Entity", "From Status", "To Status", "Trigger", "Task / Function"]
for c, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=c, value=h)
style_header(ws4, len(headers4))

transitions = [
    ("flywheel_runs", "(created)",  "PENDING",    "Job creation via API",                         "endpoints.py:create_job()"),
    ("flywheel_runs", "PENDING",    "RUNNING",    "Pipeline starts executing",                    "tasks.py:initialize_workflow()"),
    ("flywheel_runs", "RUNNING",    "COMPLETED",  "All NIMs processed, no errors",                "tasks.py:finalize_flywheel_run()"),
    ("flywheel_runs", "RUNNING",    "FAILED",     "Unrecoverable error in any stage",             "tasks.py:run_nim_workflow_dag() exception handler"),
    ("flywheel_runs", "RUNNING",    "CANCELLED",  "User cancels via API",                         "job_manager.py:cancel_job()"),
    ("", "", "", "", ""),
    ("nims",          "(created)",  "PENDING",    "initialize_workflow inserts NIM record",        "tasks.py:initialize_workflow()"),
    ("nims",          "PENDING",    "RUNNING",    "NIM deployed and model synced",                 "tasks.py:spin_up_nim()"),
    ("nims",          "RUNNING",    "COMPLETED",  "All evals done, NIM shut down",                 "tasks.py:shutdown_deployment()"),
    ("nims",          "RUNNING",    "FAILED",     "Error during eval or customization",            "tasks.py:shutdown_deployment()"),
    ("nims",          "RUNNING",    "CANCELLED",  "Cancellation detected",                         "tasks.py:shutdown_deployment()"),
    ("", "", "", "", ""),
    ("nims.deployment_status", "CREATED",  "PENDING", "DMS accepts deployment request",            "dms_client.py:deploy_model()"),
    ("nims.deployment_status", "PENDING",  "READY",   "NIM container running + model loaded",      "dms_client.py:wait_for_deployment()"),
    ("nims.deployment_status", "READY",    "COMPLETED","NIM shut down after pipeline stage",       "dms_client.py:shutdown_deployment()"),
    ("", "", "", "", ""),
    ("evaluations",   "(created)",  "IN_PROGRESS","Eval job submitted to NeMo Evaluator",          "tasks.py:run_generic_eval()"),
    ("evaluations",   "IN_PROGRESS","COMPLETED",  "NeMo Evaluator returns results",                "tasks.py:run_generic_eval()"),
    ("evaluations",   "IN_PROGRESS","FAILED",     "Evaluation error or cancellation",              "tasks.py:run_generic_eval()"),
    ("", "", "", "", ""),
    ("customizations","(created)",  "IN_PROGRESS","Training job submitted to NeMo Customizer",     "tasks.py:start_customization()"),
    ("customizations","IN_PROGRESS","COMPLETED",  "Fine-tuning done + model synced to NIM",        "tasks.py:start_customization()"),
    ("customizations","IN_PROGRESS","FAILED",     "Training error or cancellation",                "tasks.py:start_customization()"),
]

for i, row in enumerate(transitions, 2):
    for c, val in enumerate(row, 1):
        ws4.cell(row=i, column=c, value=val)
    use_font = BOLD_FONT if row[0] and not row[1] else BODY_FONT
    style_row(ws4, i, len(headers4), font=use_font)

auto_width(ws4, len(headers4))
ws4.freeze_panes = "A2"


# ════════════════════════════════════════════════════
# Sheet 5: External API Calls
# ════════════════════════════════════════════════════
ws5 = wb.create_sheet("External API Calls")
ws5.sheet_properties.tabColor = "7030A0"

headers5 = ["Service", "Method", "Endpoint", "Purpose", "Called From", "Polling?"]
for c, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=c, value=h)
style_header(ws5, len(headers5))

api_calls = [
    ("NeMo DMS",       "POST",   "/v1/deployment/model-deployments",                "Deploy NIM container",            "dms_client.py:deploy_model()",          "No"),
    ("NeMo DMS",       "GET",    "/v1/deployment/model-deployments/{ns}/{model}",    "Check deployment status",         "dms_client.py:wait_for_deployment()",    "Yes (5s)"),
    ("NeMo DMS",       "DELETE", "/v1/deployment/model-deployments/{ns}/{model}",    "Shutdown NIM container",          "dms_client.py:shutdown_deployment()",    "No"),
    ("NIM Proxy",      "GET",    "/v1/models",                                       "Check if model is loaded/synced", "dms_client.py:wait_for_model_sync()",    "Yes (30s)"),
    ("NeMo Evaluator", "POST",   "/v1/evaluation/jobs",                              "Submit evaluation job",           "evaluator.py:run_evaluation()",          "No"),
    ("NeMo Evaluator", "GET",    "/v1/evaluation/jobs/{id}",                         "Check eval status/progress",      "evaluator.py:wait_for_evaluation()",     "Yes (5s)"),
    ("NeMo Evaluator", "GET",    "/v1/evaluation/jobs/{id}/results",                 "Fetch evaluation results",        "evaluator.py:get_evaluation_results()",  "No"),
    ("NeMo Evaluator", "DELETE", "/v1/evaluation/jobs/{id}",                         "Delete evaluation job",           "evaluator.py:delete_evaluation_job()",   "No"),
    ("NeMo Customizer","POST",   "/v1/customization/jobs",                           "Submit LoRA SFT training job",    "customizer.py:start_training_job()",     "No"),
    ("NeMo Customizer","GET",    "/v1/customization/jobs/{id}",                      "Check training job status",       "customizer.py:wait_for_customization()", "Yes"),
    ("NeMo Customizer","DELETE", "/v1/customization/models/{name}",                  "Delete fine-tuned model",         "customizer.py:delete_customized_model()","No"),
    ("NeMo Datastore", "POST",   "HuggingFace datasets API",                        "Upload train/eval JSONL dataset", "data_uploader.py:upload()",              "No"),
    ("NeMo Datastore", "DELETE", "HuggingFace datasets API",                        "Delete dataset",                  "data_uploader.py:delete_dataset()",      "No"),
    ("KDB-X",          "IPC",    "SyncQConnection (port 8082)",                      "All DB reads/writes + HNSW",      "kdbx/connection.py:pykx_connection()",   "No"),
    ("Redis",          "TCP",    "redis://df-redis-service:6379/0",                  "Celery task broker + results",    "src/tasks/tasks.py (Celery config)",     "No"),
    ("MLflow",         "HTTP",   "http://df-mlflow-service:5000",                    "Experiment tracking + artifacts",  "Various eval tasks",                     "No"),
    ("NVIDIA API",     "POST",   "https://integrate.api.nvidia.com/v1/chat/completions","Remote LLM judge inference", "llm_as_judge.py", "No"),
]

for i, row in enumerate(api_calls, 2):
    for c, val in enumerate(row, 1):
        ws5.cell(row=i, column=c, value=val)
    style_row(ws5, i, len(headers5))

auto_width(ws5, len(headers5))
ws5.auto_filter.ref = "A1:F{}".format(len(api_calls) + 1)
ws5.freeze_panes = "A2"


# ════════════════════════════════════════════════════
# Sheet 6: Workflow Summary
# ════════════════════════════════════════════════════
ws6 = wb.create_sheet("Workflow Summary")
ws6.sheet_properties.tabColor = "ED7D31"
# Move to first position
wb.move_sheet(ws6, offset=-5)

headers6 = ["#", "Workflow", "Entry Point", "Components Involved", "External Services", "Description"]
for c, h in enumerate(headers6, 1):
    ws6.cell(row=1, column=c, value=h)
style_header(ws6, len(headers6))

summaries = [
    (1, "Job Creation",       "POST /api/jobs",           "FastAPI, KDBXCollection, Redis",
     "KDB-X, Redis",
     "User submits workload_id + client_id. FlywheelRun inserted into KDB-X with PENDING status. "
     "Celery task enqueued on parent_queue. Returns job ID immediately."),

    (2, "Flywheel Pipeline",  "run_nim_workflow_dag (Celery parent_queue)",
     "Parent Worker, Celery Worker, TaskDBManager, KDBXCollection, DMSClient, Evaluator, Customizer, DataUploader",
     "KDB-X, Redis, NeMo DMS, NIM, NeMo Evaluator, NeMo Customizer, NeMo Datastore, MLflow",
     "5-stage DAG: (1) Initialize - set RUNNING, create NIM records. (2) Create Datasets - read logs, "
     "optional market enrichment, split, upload to Datastore. (3) LLM Judge - deploy/validate judge model. "
     "(4) Per-NIM: spin up -> base eval -> generate signals (base) -> backtest (base) -> customization -> cust eval -> generate signals (customized) -> backtest (customized) -> shutdown. "
     "(5) Finalize - mark COMPLETED."),

    (3, "Job Cancellation",   "POST /api/jobs/{id}/cancel",
     "FastAPI, JobService, FlywheelJobManager, TaskDBManager",
     "KDB-X, Redis",
     "Sets flywheel_runs.status=CANCELLED in KDB-X (database flag). All running pipeline tasks detect "
     "cancellation on their next check_cancellation() poll. Critical stages raise immediately; NIM stages "
     "return gracefully to allow resource cleanup."),

    (4, "Job Deletion",       "DELETE /api/jobs/{id}",
     "FastAPI, JobService, FlywheelJobManager, Customizer, Evaluator, DataUploader",
     "KDB-X, NeMo Customizer, NeMo Evaluator, NeMo Datastore, MLflow",
     "Hard delete: removes fine-tuned models from Customizer, eval jobs from Evaluator, datasets from Datastore, "
     "MLflow experiments. Then cascade-deletes all KDB-X records: evaluations -> customizations -> nims -> "
     "llm_judge_runs -> flywheel_runs."),

    (5, "Vector Search",      "index_embeddings_to_es() / search_similar_embeddings()",
     "HNSW Adapter (es_adapter.py), KDBXCollection",
     "KDB-X (.ai HNSW module)",
     "Elasticsearch replacement using native KDB-X HNSW. Embeddings persisted to flywheel_embeddings table. "
     "Index built server-side via .ai.hnsw.put (cosine similarity, M=32). Search via .ai.hnsw.search returns "
     "top-k (score, tool_name, record) tuples. Indexes rebuilt from table on restart."),

    (6, "Data Explorer and Analytics", "GET /api/data/*, POST /api/backtest",
     "FastAPI, KDBXCollection",
     "KDB-X",
     "Schema introspection (meta each table), paginated table browsing (select[N] from table), record counts. "
     "Financial backtesting: runs q analytics on signals table computing Sharpe ratio, max drawdown, total return, win rate."),
]

for i, row in enumerate(summaries, 2):
    for c, val in enumerate(row, 1):
        ws6.cell(row=i, column=c, value=val)
    wf_num = row[0]
    fill = WORKFLOW_FILLS.get(wf_num)
    style_row(ws6, i, len(headers6), fill=fill)
    ws6.row_dimensions[i].height = 80

auto_width(ws6, len(headers6))
ws6.freeze_panes = "A2"


# ── Save ──
out = "/Users/abdalhamidalattar/Projects/nvidia/ai-model-distillation-for-financial-data/docs/diagrams/workflow-sequence-diagram.xlsx"
wb.save(out)
print("Saved to " + out)
